from flask import Flask, render_template, jsonify, request
import mysql.connector
import io
import os
import json
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from flask import send_file

# A password e o limiar de motor são lidos do config.py (que lê o .env).
from config import DB_CONFIG, CORRENTE_MINIMA_MOTOR

app = Flask(__name__)

def query(sql, params=None):
    conn   = mysql.connector.connect(**DB_CONFIG)
    cursor = conn.cursor(dictionary=True)
    cursor.execute(sql, params or ())
    rows = cursor.fetchall()
    cursor.close()
    conn.close()
    return rows

def query_raw(sql):
    conn   = mysql.connector.connect(**DB_CONFIG)
    cursor = conn.cursor(dictionary=True)
    cursor.execute(sql)
    rows = cursor.fetchall()
    cursor.close()
    conn.close()
    return rows

@app.route("/")
def index():
    return render_template("dashboard.html")

# ════════════════════════════════════════════════════════════════════
# ROTA SEGURA (Sem crash de % no MySQL)
# ════════════════════════════════════════════════════════════════════
@app.route("/api/leituras")
def api_leituras():
    periodo_recebido = request.args.get("periodo", "30 MINUTE")
    periodos_validos = {
        "5 MINUTE": "5 MINUTE", "15 MINUTE": "15 MINUTE", "30 MINUTE": "30 MINUTE",
        "1 HOUR": "1 HOUR", "1 DAY": "1 DAY", "7 DAY": "7 DAY"
    }
    intervalo = periodos_validos.get(periodo_recebido, "30 MINUTE")

    # Devolve a data completa sempre. O Frontend trata de deixar bonito.
    dados = query_raw(f"""
        SELECT topico, temperatura, rms_mm_s, corrente_A,
               DATE_FORMAT(criado_em, '%Y-%m-%d %H:%i:%s') AS hora
        FROM leituras_sensores
        WHERE criado_em >= NOW() - INTERVAL {intervalo}
        ORDER BY criado_em ASC
    """)
    return jsonify(dados)

CAMINHO_LIMITES = os.path.join(os.path.dirname(os.path.abspath(__file__)), "limites_fase1.json")

@app.route("/api/limites-controlo")
def api_limites_controlo():
    if not os.path.exists(CAMINHO_LIMITES):
        return jsonify({"disponivel": False, "erro": "Ficheiro JSON nao encontrado"})

    try:
        with open(CAMINHO_LIMITES, encoding="utf-8") as f:
            dados = json.load(f)
            
        if "limites" in dados:
            return jsonify({"disponivel": True, "limites": dados["limites"]})
        else:
            return jsonify({"disponivel": False, "erro": "Formato JSON incorreto"})
    except Exception as e:
        return jsonify({"disponivel": False, "erro": str(e)})

@app.route("/api/alertas")
def api_alertas():
    dados = query_raw("""
        SELECT id, device_id, topico, nivel, mensagem, valor,
               DATE_FORMAT(criado_em, '%H:%i:%s') AS hora, resolvido
        FROM alertas WHERE resolvido = FALSE ORDER BY criado_em DESC LIMIT 20
    """)
    return jsonify(dados)

@app.route("/api/resumo")
def api_resumo():
    dados = query("""
        SELECT
            (SELECT COUNT(*) FROM leituras_sensores WHERE criado_em >= NOW() - INTERVAL 1 MINUTE) AS leituras_ultimo_min,
            (SELECT COUNT(*) FROM alertas WHERE resolvido = FALSE) AS alertas_ativos,
            (SELECT temperatura FROM leituras_sensores WHERE temperatura IS NOT NULL ORDER BY criado_em DESC LIMIT 1) AS temp_inst,
            (SELECT corrente_A FROM leituras_sensores WHERE corrente_A IS NOT NULL ORDER BY criado_em DESC LIMIT 1) AS corrente_inst,
            (SELECT rms_mm_s FROM leituras_sensores WHERE rms_mm_s IS NOT NULL ORDER BY criado_em DESC LIMIT 1) AS vib_inst
    """)
    return jsonify(dados[0] if dados else {})

@app.route("/api/estado-motor")
def api_estado_motor():
    dados = query("""
        SELECT corrente_A, DATE_FORMAT(criado_em, '%H:%i:%s') AS hora,
               TIMESTAMPDIFF(SECOND, criado_em, NOW()) AS segundos
        FROM leituras_sensores
        WHERE topico = 'sensor/corrente' AND corrente_A IS NOT NULL
        ORDER BY criado_em DESC LIMIT 1
    """)

    if not dados:
        return jsonify({"ligado": False, "corrente": None, "limiar": CORRENTE_MINIMA_MOTOR, "motivo": "sem_dados"})

    corrente = float(dados[0]["corrente_A"])
    segundos = dados[0]["segundos"]

    if segundos is not None and segundos > 30:
        return jsonify({"ligado": False, "corrente": round(corrente, 3), "limiar": CORRENTE_MINIMA_MOTOR, "motivo": "leitura_antiga"})

    return jsonify({"ligado": corrente >= CORRENTE_MINIMA_MOTOR, "corrente": round(corrente, 3), "limiar": CORRENTE_MINIMA_MOTOR, "motivo": "ok"})

@app.route("/api/estatisticas")
def api_estatisticas():
    periodo = request.args.get("periodo", "1 HOUR")
    periodos_validos = { "5 MINUTE": "Últimos 5 minutos", "15 MINUTE": "Últimos 15 minutos", "30 MINUTE": "Últimos 30 minutos", "1 HOUR": "Última hora", "1 DAY": "Último dia", "7 DAY": "Última semana" }
    if periodo not in periodos_validos: periodo = "1 HOUR"

    dados = query(f"""
        SELECT topico, COUNT(*) AS total_leituras,
            ROUND(MIN(temperatura), 2) AS temp_min, ROUND(MAX(temperatura), 2) AS temp_max, ROUND(AVG(temperatura), 2) AS temp_media, ROUND(STDDEV(temperatura), 3) AS temp_desvio,
            ROUND(MIN(rms_mm_s), 3) AS vib_min, ROUND(MAX(rms_mm_s), 3) AS vib_max, ROUND(AVG(rms_mm_s), 3) AS vib_media, ROUND(STDDEV(rms_mm_s), 4) AS vib_desvio,
            ROUND(MIN(corrente_A), 3) AS corr_min, ROUND(MAX(corrente_A), 3) AS corr_max, ROUND(AVG(corrente_A), 3) AS corr_media, ROUND(STDDEV(corrente_A), 4) AS corr_desvio
        FROM leituras_sensores
        WHERE criado_em >= NOW() - INTERVAL {periodo}
        GROUP BY topico ORDER BY topico
    """)
    return jsonify({"periodo": periodos_validos[periodo], "sensores": dados})

@app.route("/api/saude")
def api_saude():
    sensores = [{"topico": "sensor/temperatura", "nome": "Temperatura"}, {"topico": "sensor/vibracao", "nome": "Vibracao"}, {"topico": "sensor/corrente", "nome": "Corrente"}]
    resultado = []
    for s in sensores:
        rows = query_raw(f"SELECT criado_em, TIMESTAMPDIFF(SECOND, criado_em, NOW()) AS segundos FROM leituras_sensores WHERE topico = '{s['topico']}' ORDER BY criado_em DESC LIMIT 1")
        if rows:
            segundos = rows[0]["segundos"]
            estado = "online" if segundos <= 10 else ("lento" if segundos <= 30 else "offline")
            ultima = str(rows[0]["criado_em"])
        else:
            estado, ultima = "offline", "Nunca"
        resultado.append({"topico": s["topico"], "nome": s["nome"], "estado": estado, "ultima": ultima})
    return jsonify(resultado)

@app.route("/api/alertas/resolver/<int:alerta_id>", methods=["POST"])
def resolver_alerta(alerta_id):
    try:
        conn = mysql.connector.connect(**DB_CONFIG)
        cursor = conn.cursor()
        cursor.execute("UPDATE alertas SET resolvido = TRUE, resolvido_em = NOW(), resolvido_por = 'operador' WHERE id = %s", (alerta_id,))
        conn.commit()
        cursor.close(); conn.close()
        return jsonify({"ok": True})
    except Exception as e: return jsonify({"ok": False, "erro": str(e)}), 500

@app.route("/api/alertas/resolver-todos", methods=["POST"])
def resolver_todos_alertas():
    try:
        conn = mysql.connector.connect(**DB_CONFIG)
        cursor = conn.cursor()
        cursor.execute("UPDATE alertas SET resolvido = TRUE, resolvido_em = NOW(), resolvido_por = 'operador' WHERE resolvido = FALSE")
        conn.commit()
        cursor.close(); conn.close()
        return jsonify({"ok": True})
    except Exception as e: return jsonify({"ok": False, "erro": str(e)}), 500

@app.route("/exportar-excel")
def exportar_excel():
    periodo = request.args.get("periodo", "5 MINUTE")
    if periodo not in ["5 MINUTE", "15 MINUTE", "30 MINUTE", "1 HOUR", "1 DAY", "7 DAY"]: periodo = "5 MINUTE"

    wb = Workbook()
    wb.remove(wb.active)

    def estilo_cabecalho(cor_hex): return {"fill": PatternFill("solid", fgColor=cor_hex), "font": Font(bold=True, color="FFFFFF", size=11), "align": Alignment(horizontal="center", vertical="center"), "border": Border(bottom=Side(style="thin", color="FFFFFF"))}
    def aplicar_cabecalho(ws, colunas, cor):
        est = estilo_cabecalho(cor)
        for col, titulo in enumerate(colunas, 1):
            c = ws.cell(row=1, column=col, value=titulo)
            c.fill, c.font, c.alignment = est["fill"], est["font"], est["align"]
            ws.column_dimensions[get_column_letter(col)].width = max(len(titulo) + 4, 14)

    SENSORES_EXCEL = [{"topico": "sensor/temperatura", "nome": "Temperatura", "campo": "temperatura", "unidade": "°C", "cor": "1D9E75"}, {"topico": "sensor/vibracao", "nome": "Vibracao", "campo": "rms_mm_s", "unidade": "mm/s", "cor": "534AB7"}, {"topico": "sensor/corrente", "nome": "Corrente", "campo": "corrente_A", "unidade": "A", "cor": "BA7517"}]
    nomes_periodo = {"5 MINUTE": "Ultimos 5 minutos", "15 MINUTE": "Ultimos 15 minutos", "30 MINUTE": "Ultimos 30 minutos", "1 HOUR": "Ultima hora", "1 DAY": "Ultimo dia", "7 DAY": "Ultima semana"}

    def cartas_controlo(valores, i, janela=20, k=3, truncar_zero=False):
        bloco = valores[max(0, i - janela + 1):i + 1]
        n = len(bloco)
        if n < 2: return {"lc": None, "lsc": None, "lic": None, "sigma": None, "z": None, "fora": None}
        media = sum(bloco) / n
        sigma = (sum((v - media) ** 2 for v in bloco) / n) ** 0.5
        lsc, lic = media + k * sigma, media - k * sigma
        if truncar_zero and lic < 0: lic = 0.0
        valor = valores[i]
        fora = 1 if (valor > lsc or valor < lic) else 0
        return {"lc": round(media, 3), "lsc": round(lsc, 3), "lic": round(lic, 3), "sigma": round(sigma, 4), "z": round((valor - media) / sigma if sigma > 0 else 0.0, 3), "fora": fora}

    fill_normal, fill_alerta, fill_fora = PatternFill("solid", fgColor="F5F5F5"), PatternFill("solid", fgColor="FFF3CD"), PatternFill("solid", fgColor="FFDADA")
    periodo_pesado = periodo in ["1 DAY", "7 DAY"]

    for sensor in SENSORES_EXCEL:
        ws = wb.create_sheet(title=sensor["nome"])
        campo = sensor["campo"]
        leituras = query_raw(f"SELECT device_id, {campo} AS valor, status, DATE_FORMAT(criado_em, '%Y-%m-%d %H:%i:%s') AS timestamp FROM leituras_sensores WHERE topico = '{sensor['topico']}' AND criado_em >= NOW() - INTERVAL {periodo} AND {campo} IS NOT NULL ORDER BY criado_em ASC")
        aplicar_cabecalho(ws, ["Timestamp", "Device ID", f"Valor ({sensor['unidade']})", "LC", "LSC", "LIC", "Sigma", "Z-score", "Fora_controlo", "Status"], sensor["cor"])
        valores = [float(r["valor"]) for r in leituras]

        for i, row in enumerate(leituras):
            if periodo_pesado:
                linha = [row["timestamp"], row["device_id"], float(row["valor"]), "-", "-", "-", "-", "-", "-", row["status"]]
                ws.append(linha)
            else:
                cc = cartas_controlo(valores, i, truncar_zero=(campo == "corrente_A"))
                linha = [row["timestamp"], row["device_id"], float(row["valor"]), cc["lc"], cc["lsc"], cc["lic"], cc["sigma"], cc["z"], cc["fora"], row["status"]]
                ws.append(linha)
                if cc["fora"] == 1 or row["status"] != "normal":
                    for col in range(1, len(linha) + 1): ws.cell(row=i + 2, column=col).fill = (fill_fora if cc["fora"] == 1 else fill_alerta)

    ws_resumo = wb.create_sheet(title="Resumo", index=0)
    aplicar_cabecalho(ws_resumo, ["Sensor", "Unidade", "Total leituras", "Minimo", "Maximo", "Media", "Desvio padrao", "Periodo"], "2C2C2A")

    for i, sensor in enumerate(SENSORES_EXCEL):
        campo = sensor["campo"]
        stats = query_raw(f"SELECT COUNT(*) AS n, ROUND(MIN({campo}), 3) AS minimo, ROUND(MAX({campo}), 3) AS maximo, ROUND(AVG({campo}), 3) AS media, ROUND(STDDEV({campo}), 4) AS desvio FROM leituras_sensores WHERE topico = '{sensor['topico']}' AND criado_em >= NOW() - INTERVAL {periodo} AND {campo} IS NOT NULL")
        s = stats[0] if stats else {}
        ws_resumo.append([sensor["nome"], sensor["unidade"], s.get("n"), s.get("minimo"), s.get("maximo"), s.get("media"), s.get("desvio"), nomes_periodo.get(periodo)])
        for col in range(1, 8): ws_resumo.cell(row=i + 2, column=col).alignment = Alignment(horizontal="center")

    ws_alertas = wb.create_sheet(title="Alertas")
    aplicar_cabecalho(ws_alertas, ["Timestamp", "Device ID", "Topico", "Nivel", "Mensagem", "Valor", "Limiar"], "A32D2D")
    alertas = query_raw("SELECT DATE_FORMAT(criado_em, '%Y-%m-%d %H:%i:%s') AS timestamp, device_id, topico, nivel, mensagem, valor, limiar FROM alertas ORDER BY criado_em DESC LIMIT 200")
    for row in alertas:
        ws_alertas.append([row["timestamp"], row["device_id"], row["topico"], row["nivel"], row["mensagem"], row["valor"], row["limiar"]])
        for col in range(1, 8):
            c = ws_alertas.cell(row=ws_alertas.max_row, column=col)
            c.fill = PatternFill("solid", fgColor="FFDADA") if row["nivel"] == "CRITICO" else PatternFill("solid", fgColor="FFF3CD")
            c.alignment = Alignment(horizontal="center")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(buffer, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name=f"iot_sensores_{periodo.replace(' ', '_').lower()}.xlsx")

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=False)